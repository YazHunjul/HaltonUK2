import io
from Programs import genInfo as GI
from Programs import canopy
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image
import streamlit as st
from PIL import Image as PILImage
from openpyxl.worksheet.header_footer import HeaderFooter
import math
from openpyxl.styles import Font
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import email.utils

# Helper function to format flow rates consistently
def format_flowrate(value, decimal_places=3):
    """Format a flowrate value to the specified number of decimal places.
    
    Args:
        value: The flowrate value to format
        decimal_places: Number of decimal places to round to (default: 3)
        
    Returns:
        Formatted flowrate as a string with exact decimal places
    """
    if value is None:
        return ""
    
    format_str = f"{{:.{decimal_places}f}}"
    return format_str.format(float(value))

def TandC():
    GI.titleAndLogo("Testing And Commissioning")
    
    # Only load URL parameters once per session
    if 'url_params_loaded' not in st.session_state:
        try:
            GI.load_from_query_params()
            # Mark that we've loaded the URL parameters
            st.session_state['url_params_loaded'] = True
        except Exception as e:
            st.error(f"Error loading form data: {e}")
    
    genInfo = GI.getGenInfo()
    
    # Create sidebar for navigation
    with st.sidebar:
        st.header("Navigation")
        st.markdown("#### Jump to canopy:")
        
        # Create anchor for top of page
        st.markdown('<a name="top"></a>', unsafe_allow_html=True)
        
        # Add link to top of page
        st.markdown('[General Information](#top)', unsafe_allow_html=True)
    
    # Continue with the main form elements
    hoods, edge_box_details = canopy.numCanopies()  # Unpack returned values
    
    # Update sidebar with canopy links after we have the hood data
    with st.sidebar:
        for key, hood in hoods.items():
            # Extract the number from the key since hood.id might not exist
            # Key format is '{selection} {num}'
            hood_num = key.split()[-1]  # Get the last part which should be the number
            
            # Create unique anchor ID for each canopy
            anchor_id = f"canopy_{hood_num}"
            
            # Format the link to show model (drawing number) - location
            drawing_info = hood.drawingNum if hood.drawingNum else f"#{int(hood_num)+1}"
            st.markdown(f'[{hood.model} ({drawing_info}) - {hood.location}](#{anchor_id})', unsafe_allow_html=True)
        
        # Add links to other sections
        st.markdown("#### Other Sections:")
        st.markdown('[Comments](#comments)', unsafe_allow_html=True)
        st.markdown('[Signature](#signature)', unsafe_allow_html=True)
    
    # Get comments BEFORE signature
    st.markdown('<a name="comments"></a>', unsafe_allow_html=True)
    st.markdown("### Comments")
    comments = GI.get_comments()
    
    # Store the signature data to prevent it from disappearing on re-renders
    if 'signature_data' not in st.session_state:
        st.session_state.signature_data = None
    
    # Get the signature input AFTER comments
    st.markdown('<a name="signature"></a>', unsafe_allow_html=True)
    st.markdown("### Signature")
    
    # IMPORTANT: Always ensure there's a valid signature_data in session state
    sign = GI.get_sign()
    
    # Save the signature data to session state if it exists
    if sign.image_data is not None:
        st.session_state.signature_data = sign.image_data
    
    # Add sharing options at the bottom
    st.markdown("---")
    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button("Save to Excel"):
            saveToExcel(genInfo, hoods, comments, sign, edge_box_details)
    with col2:
        with st.expander("📤 Share Your Form Data To On-Site Technician", expanded=False):
            st.write("Generate a link you can share with others. When they open the link, all your form data will be pre-filled.")
            GI.display_shareable_link()
        
def saveToExcel(genInfo, hoods, comments, sign, edge_box_details):
    wb = openpyxl.load_workbook('T&C_Templates.xlsx')
    ws = wb.active
    
    #General Information
    row=4
    genFont(ws, 'A', row, f"CLIENT: {genInfo['client'].title()} ")
    row+=2
    genFont(ws, 'A', row, f"PROJECT NAME: {genInfo['project_name'].title()} ")
    row+=2
    genFont(ws, 'A', row, f"PROJECT NUMBER: {genInfo['project_number'].title()} ")
    row+=2
    genFont(ws, 'A', row, f"DATE OF VISIT: {genInfo['date_of_visit']} ")
    row+=2
    genFont(ws, 'A', row, f"ENGINEER(s): {genInfo['engineers'].title()} ")
    row+=4
    # Apply the fill and border dynamically to a range of cells
    if canopy.features and len(canopy.features) > 0:
        genFont(ws, 'A', row, "Your Kitchen Ventilation System has the following serviceable technology:")
        ws[f'A{row}'].font = Font(bold=True, size=15)
        row += 3

        # Display UV Image if Selected
        if "CJ" in canopy.features and canopy.features["CJ"] == True:
            original_image = PILImage.open("features/C-RAY.png")
            resized_image = original_image.resize((600, 150))  # Resize the image
            resized_image.save("resized_image.png")
            img = Image("resized_image.png")
            ws.add_image(img, f'A{row}')
            row += 8  # Adjust the row position after the image

        # Display M.A.R.V.E.L. Image if Selected
        if "marvel" in canopy.features and canopy.features["marvel"] == True:
            original_image = PILImage.open("features/Marvel.png")
            resized_image = original_image.resize((620, 130))  # Resize the image
            resized_image.save("resized_Marvel.png")
            img = Image("resized_Marvel.png")
            ws.add_image(img, f'A{row}')
            row += 5  # Adjust the row position after the image
    row+=10
    ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=row))
    row+=2
    #Canopy Air Readings
    titleFont(ws, 'A', row, "KITCHEN CANOPY AIR READINGS")
    row+=2
    
    fill_style = PatternFill(start_color="9ac9f4", end_color="9ac9f4", fill_type="solid")
# Define a border (thin border for all sides)
    thin_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
# Should be fixed here
    # st.write(hoods.items())
    for k,v in hoods.items():
        #Capture Jet Hoods
        if v.model in ['KVF', 'KVI', 'KCH-F', 'KCH-I', 'KSR-S', 'KSR-F', 'KSR-M', 'UVF', 'UVI', 'USR-S', 'USR-F', 'USR-M', 'UWF'] or (v.model in ['CMW-FMOD', 'CMW-IMOD'] and 'slot_length' not in v.sections):
            colorFill(ws, row)
            ws[f'A{row}'].border = Border(top=Side(style='thin'),
            bottom=Side(style='thin'))
            #ws[f'j{row}'].border = Border(left=Side(style='thin'))
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f'A{row}:I{row}')
            colorFill(ws, row)  # Apply the fill and border to the entire range
            genFont(ws, 'A', row, "EXTRACT AIR DATA", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Basic hood information in CXW style format
            genFont(ws, 'A', row, f'Drawing Number')
            genFont(ws, 'C', row, f'{v.drawingNum}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Location')
            genFont(ws, 'C', row, f'{v.location}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Model')
            genFont(ws, 'C', row, f'{v.model}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Design Flowrate (m³/s)')
            genFont(ws, 'C', row, f'{v.total_design_flow_ms}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Quantity of Sections')
            genFont(ws, 'C', row, f'{v.quantityOfSections}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Calculation')
            genFont(ws, 'C', row, f'QV = Kf x √Pa')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # Add With M.A.R.V.E.L. row
            marvel_status = "Yes" if "M.A.R.V.E.L. System" in v.checklist else "No"
            genFont(ws, 'A', row, f'With M.A.R.V.E.L.')
            genFont(ws, 'C', row, f'{marvel_status}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR READINGS", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            extendRow(ws, row)
            # Merge A and B for Section #
            ws.merge_cells(f'A{row}:B{row}')
            genFont(ws, 'A', row, 'Section #')
            makeCenter(ws, 'A', row)
            sectionBorder(ws, row, 1, 3)

            # TAB Reading - now in column C
            genFont(ws, 'C', row, 'T.A.B Point Reading (Pa)')
            makeCenter(ws, 'C', row)
            sectionBorder(ws, row, 3, 4)

            # K-Factor - now in column D
            genFont(ws, 'D', row, 'K-Factor (m³/h)')
            makeCenter(ws, 'D', row)
            sectionBorder(ws, row, 4, 5)
            
            # Flowrate m³/h - now in column E
            genFont(ws, 'E', row, 'Flowrate (m³/h)')
            makeCenter(ws, 'E', row)
            sectionBorder(ws, row, 5, 6)
            
            # Flowrate m³/s - now in column F
            genFont(ws, 'F', row, 'Flowrate (m³/s)')
            makeCenter(ws, 'F', row)
            sectionBorder(ws, row, 6, 7)
            
            # Min % - now in column G
            genFont(ws, 'G', row, 'Min (%)')
            makeCenter(ws, 'G', row)
            sectionBorder(ws, row, 7, 8)
            
            # Idle % - now in column H
            genFont(ws, 'H', row, 'Idle (%)')
            makeCenter(ws, 'H', row)
            sectionBorder(ws, row, 8, 9)
            
            # Design m³/s - now in column I
            genFont(ws, 'I', row, 'Design (m³/s)')
            makeCenter(ws, 'I', row)
            sectionBorder(ws, row, 9, 10)
            
            row+=1
            
            # Now Fill the Info (v.sections)
            totalFlowRate = 0
            percentage = 0
            totalAchieved = 0
            total_flowrate_m3s = 0
          
            for section, info in v.sections.items():
                
                if isinstance(info, dict):  # Make sure info is a dictionary before accessing
                    # Calculate flowrate based on hood type
                    # `st.write(v.model)
                    if v.model == 'CXW':
                        if 'flowrate_m3s' in info:
                            flowrate_m3s = info['flowrate_m3s']
                    else:
                        if 'k_factor' in info and 'tab_reading' in info:
                            flowrate_m3s = round((info['k_factor'] * math.sqrt(info['tab_reading'])) / 3600, 3)
                        else:
                            flowrate_m3s = 0
                    
                    total_flowrate_m3s += flowrate_m3s
                    
                    # Merge A and B for section number
                    ws.merge_cells(f'A{row}:B{row}')
                    genFont(ws, 'A', row, section)
                    makeCenter(ws, 'A', row)
                    sectionBorder(ws, row, 1, 3)
                    
                    # TAB Reading - now in column C
                    genFont(ws, 'C', row, f' {round(info.get("tab_reading", 0), 0)} Pa')
                    makeCenter(ws, 'C', row)
                    sectionBorder(ws, row, 3, 4)
                    
                    # K-Factor - now in column D
                    genFont(ws, 'D', row, f' {info.get("k_factor", 0)}')
                    makeCenter(ws, 'D', row)
                    sectionBorder(ws, row, 4, 5)
                    
                    # AchievedFlow - now in column E
                    achieved_flow = round(info.get("achieved", 0), 0)
                    genFont(ws, 'E', row, f' {achieved_flow}')
                    makeCenter(ws, 'E', row)
                    sectionBorder(ws, row, 5, 6)
                    totalAchieved += round(info.get('achieved', 0), 2)
                    
                    # Design Flow - now in column F
                    genFont(ws, 'F', row, f' {format_flowrate(flowrate_m3s, 3)}')
                    makeCenter(ws, 'F', row)
                    sectionBorder(ws, row, 6, 7)
                    totalFlowRate += flowrate_m3s
                    
                    # Min % - now in column G
                    min_percentage = 'NA'
                    if 'M.A.R.V.E.L. System' in v.checklist and 'extract_min_pct' in info:
                        min_percentage = info['extract_min_pct']
                    genFont(ws, 'G', row, f' {min_percentage}%' if min_percentage != 'NA' else ' -')
                    makeCenter(ws, 'G', row)
                    sectionBorder(ws, row, 7, 8)
                    
                    # Idle % - now in column H
                    idle_percentage = 'NA'
                    if 'M.A.R.V.E.L. System' in v.checklist and 'extract_idle_pct' in info:
                        idle_percentage = info['extract_idle_pct']
                    genFont(ws, 'H', row, f' {idle_percentage}%' if idle_percentage != 'NA' else ' -')
                    makeCenter(ws, 'H', row)
                    sectionBorder(ws, row, 8, 9)
                    
                    # Design m³/s - now in column I
                    design_flow_per_section = 'NA'
                    # v.total_design_flow_ms / v.quantityOfSections if v.quantityOfSections > 0 else 0
                    if 'M.A.R.V.E.L. System' in v.checklist and 'extract_design_flow' in info:
                        design_flow_per_section = info['extract_design_flow']
                    genFont(ws, 'I', row, f' {round(design_flow_per_section, 2)}' if design_flow_per_section != 'NA' else ' -')
                    makeCenter(ws, 'I', row)
                    sectionBorder(ws, row, 9, 10)
                    
                    row += 1
            
            # Calculate total flowrate for all sections first
            total_flowrate_m3s = 0
            for section, info in v.sections.items():
                if isinstance(info, dict) and "k_factor" in info and "tab_reading" in info:
                    flowrate_m3s = round((info["k_factor"] * math.sqrt(info["tab_reading"])) / 3600, 3)
                    total_flowrate_m3s += flowrate_m3s

            # Format total with exact 3 decimal places
            formatted_total = format_flowrate(total_flowrate_m3s, 3)
            
            colorFill2(ws, row)
            row+=1
            # Show total flowrate with correct formatting
            ws.merge_cells(f'A{row}:C{row}')  # Merge A to C
            genFont(ws, 'A', row, f'Total Flowrate: {formatted_total} m³/s')
            makeCenter(ws, 'A', row)
            # Only add borders for the merged cell A:C
            sectionBorder(ws, row, 1, 4)
            
            row+=2
            
            # Add supply air data section for supply-capable hoods
            if v.model in ['KVF', 'KCH-F', 'UVF', 'USR-F', 'KSR-F', 'KWF', 'UWF'] or (v.model == 'CMW-FMOD' and 'slot_length' not in v.sections):
                #Supply Air Readings
                row +=2
                colorFill(ws, row)
                ws.merge_cells(f'A{row}:I{row}')
                genFont(ws, 'A', row, "SUPPLY AIR DATA", "FFFFFF")
                makeCenter(ws, 'A', row)
                row+=1

                # Basic supply hood information in CXW style format
                genFont(ws, 'A', row, f'Drawing Number')
                genFont(ws, 'C', row, f'{v.drawingNum}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                genFont(ws, 'A', row, f'Location')
                genFont(ws, 'C', row, f'{v.location}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                genFont(ws, 'A', row, f'Model')
                genFont(ws, 'C', row, f'{v.model}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                genFont(ws, 'A', row, f'Design Airflow (m³/s)')
                genFont(ws, 'C', row, f'{v.total_supply_design_flow_ms}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                genFont(ws, 'A', row, f'Quantity of Sections')
                genFont(ws, 'C', row, f'{v.quantityOfSections}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                genFont(ws, 'A', row, f'Calculation')
                genFont(ws, 'C', row, f'QV = Kf x √Pa')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                # Add With M.A.R.V.E.L. row
                marvel_status = "Yes" if "M.A.R.V.E.L. System" in v.checklist else "No"
                genFont(ws, 'A', row, f'With M.A.R.V.E.L.')
                genFont(ws, 'C', row, f'{marvel_status}')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
                
                colorFill(ws, row)
                ws.merge_cells(f'A{row}:I{row}')
                genFont(ws, 'A', row, "SUPPLY AIR READINGS", "FFFFFF")
                makeCenter(ws, 'A', row)
                row+=1
                extendRow(ws, row)
                # Merge A and B for Section #
                ws.merge_cells(f'A{row}:B{row}')
                genFont(ws, 'A', row, 'Section #')
                makeCenter(ws, 'A', row)
                sectionBorder(ws, row, 1, 3)
                
                # TAB Reading - now in column C
                genFont(ws, 'C', row, 'T.A.B Point Reading (Pa)')
                makeCenter(ws, 'C', row)
                sectionBorder(ws, row, 3, 4)
                
                # K-Factor - now in column D
                genFont(ws, 'D', row, 'K-Factor (m³/h)')
                makeCenter(ws, 'D', row)
                sectionBorder(ws, row, 4, 5)
                
                # Flowrate m³/h - now in column E
                genFont(ws, 'E', row, 'Flowrate (m³/h)')
                makeCenter(ws, 'E', row)
                sectionBorder(ws, row, 5, 6)
                
                # Flowrate m³/s - now in column F
                genFont(ws, 'F', row, 'Flowrate (m³/s)')
                makeCenter(ws, 'F', row)
                sectionBorder(ws, row, 6, 7)
                
                # Min % - now in column G
                genFont(ws, 'G', row, 'Min (%)')
                makeCenter(ws, 'G', row)
                sectionBorder(ws, row, 7, 8)
                
                # Idle % - now in column H
                genFont(ws, 'H', row, 'Idle (%)')
                makeCenter(ws, 'H', row)
                sectionBorder(ws, row, 8, 9)
                
                # Design m³/s - now in column I
                genFont(ws, 'I', row, 'Design (m³/s)')
                makeCenter(ws, 'I', row)
                sectionBorder(ws, row, 9, 10)
                
                row+=1
                
                totalFlowRateSup = 0
                totPercentage = 0
                for section2, info2 in v.sections.items():
                    if isinstance(info2, dict):  # Make sure info2 is a dictionary before accessing
                        # Merge A and B for section number
                        ws.merge_cells(f'A{row}:B{row}')
                        genFont(ws, 'A', row, section2)
                        makeCenter(ws, 'A', row)
                        sectionBorder(ws, row, 1, 3)
                        
                        # TAB Reading - now in column C
                        genFont(ws, 'C', row, f' {round(info2.get("supplyTab", 0), 0)} Pa')
                        makeCenter(ws, 'C', row)
                        sectionBorder(ws, row, 3, 4)
                        
                        # K-Factor - now in column D
                        genFont(ws, 'D', row, f' {info2.get("supplyKFactor", 0)}')
                        makeCenter(ws, 'D', row)
                        sectionBorder(ws, row, 4, 5)
                        
                        # AchievedFlow - now in column E
                        genFont(ws, 'E', row, f' {round(info2.get("achievedSupply", 0), 2)}')
                        makeCenter(ws, 'E', row)
                        sectionBorder(ws, row, 5, 6)
                        
                        # Design Flow - now in column F
                        flow_value_m3s = (info2.get("supplyKFactor", 0) * math.sqrt(info2.get("supplyTab", 0))) / 3600
                        genFont(ws, 'F', row, f' {format_flowrate(flow_value_m3s)}')
                        makeCenter(ws, 'F', row)
                        sectionBorder(ws, row, 6, 7)
                        totalFlowRateSup += flow_value_m3s
                        
                        # Min % - now in column G
                        min_percentage = 'NA'
                        if 'M.A.R.V.E.L. System' in v.checklist and 'supply_min_pct' in info2:
                            min_percentage = info2['supply_min_pct']
                        genFont(ws, 'G', row, f' {min_percentage}%' if min_percentage != 'NA' else ' -')
                        makeCenter(ws, 'G', row)
                        sectionBorder(ws, row, 7, 8)
                        
                        # Idle % - now in column H
                        idle_percentage = 'NA'
                        if 'M.A.R.V.E.L. System' in v.checklist and 'supply_idle_pct' in info2:
                            idle_percentage = info2['supply_idle_pct']
                        genFont(ws, 'H', row, f' {idle_percentage}%' if idle_percentage != 'NA' else ' -')
                        makeCenter(ws, 'H', row)
                        sectionBorder(ws, row, 8, 9)
                        
                        # Design m³/s - now in column I
                        supply_design_flow_per_section = 'NA'
                        # v.total_supply_design_flow_ms / v.quantityOfSections if v.quantityOfSections > 0 else 0
                        if 'M.A.R.V.E.L. System' in v.checklist and 'supply_design_flow' in info2:
                            supply_design_flow_per_section = info2['supply_design_flow']
                        genFont(ws, 'I', row, f' {round(supply_design_flow_per_section, 2)}' if supply_design_flow_per_section != 'NA' else ' -')
                        makeCenter(ws, 'I', row)
                        sectionBorder(ws, row, 9, 10)
                        
                        row+=1
                
                # Format total with exact 3 decimal places
                formatted_total = format_flowrate(totalFlowRateSup, 3)
                
                colorFill2(ws, row)
                row+=1
                # Show total flowrate with correct formatting
                ws.merge_cells(f'A{row}:C{row}')  # Merge A to C
                genFont(ws, 'A', row, f'Total Flowrate: {formatted_total} m³/s')
                makeCenter(ws, 'A', row)
                # Only add borders for the merged cell A:C
                sectionBorder(ws, row, 1, 4)
                
                row+=2
            
            row+=1
            #Start of Result summary (Extract Air)
        #CheckList

        if "UV Capture Jet" in v.checklist:
            row += 2
            # Add color-filled title for UV section
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, f"UV CAPTURE RAY SYSTEM FOR {v.model}", "FFFFFF")
            makeCenter(ws, 'A', row)
            row += 1

            # Add checklist items
            for check, value in v.checklist["UV Capture Jet"].items():
                # Add description in A:G
                ws.merge_cells(f'A{row}:G{row}')
                genFont(ws, 'A', row, check)
                
                # Add value in H:I
                ws.merge_cells(f'H{row}:I{row}')
                if value is True:
                    ws[f'H{row}'] = "✓"
                elif value is False:
                    ws[f'H{row}'] = ""
                else:
                    ws[f'H{row}'] = value
                makeCenter(ws, 'H', row)

                # Apply borders to H and I
                border = Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
                ws[f'H{row}'].border = border
                ws[f'I{row}'].border = border

                # Apply borders to the entire row
                sectionBorder(ws, row, 1, 9)
                row += 1

        # Add M.A.R.V.E.L. details if available
        if "M.A.R.V.E.L. System" in v.checklist:
            # row += 2
            # # Add color-filled title for Marvel section
            # colorFill(ws, row)
            # ws.merge_cells(f'A{row}:I{row}')
            # genFont(ws, 'A', row, f"M.A.R.V.E.L. SYSTEM FOR {v.model}", "FFFFFF")
            # makeCenter(ws, 'A', row)
            # row += 1

            # Add checklist items
            for check, value in v.checklist["M.A.R.V.E.L. System"].items():
                # Add description in A:G
                ws.merge_cells(f'A{row}:G{row}')
                genFont(ws, 'A', row, check)
                
                # Add value in H:I
                ws.merge_cells(f'H{row}:I{row}')
                if value is True:
                    ws[f'H{row}'] = "✓"
                elif value is False:
                    ws[f'H{row}'] = ""
                else:
                    ws[f'H{row}'] = value
                makeCenter(ws, 'H', row)

                # Apply borders to H and I
                border = Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
                ws[f'H{row}'].border = border
                ws[f'I{row}'].border = border

                # Apply borders to the entire row
                sectionBorder(ws, row, 1, 9)
                row += 1

        # Changed 'elif' to 'if' to process CXW hoods regardless of M.A.R.V.E.L. status
        if v.model == 'CXW':
            st.write(v)
            # Add EXTRACT AIR DATA header
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR DATA", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Basic hood information
            genFont(ws, 'A', row,f'Drawing Number')
            genFont(ws, 'C', row,f'{v.drawingNum}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Canopy Location')
            genFont(ws, 'C', row, f'{v.location}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Canopy Model')
            genFont(ws, 'C', row, f'CXW')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Design Flowrate (m³/s)')
            genFont(ws, 'C', row, f'{v.total_design_flow_ms}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Quantity of Grills')
            genFont(ws, 'C', row, f'{v.quantityOfSections}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # Add Grill Size row after Quantity of Grills
            if v.sections and '1' in v.sections:  # Get grill size from first section
                grill_size = v.sections['1'].get('grill_size', '')
                genFont(ws, 'A', row, f'Grill Size')
                genFont(ws, 'C', row, f'{grill_size} mm')
                ws.merge_cells(f'A{row}:B{row}')
                ws.merge_cells(f'C{row}:I{row}')
                sectionBorder(ws, row, 1, 3)
                sectionBorder(ws, row, 3, 10)
                row+=1
            
            genFont(ws, 'A', row, f'Calculation')
            genFont(ws, 'C', row, f'Qv = A x v')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1

            # Add M.A.R.V.E.L. status row
            marvel_status = "Yes" if "M.A.R.V.E.L. System" in v.checklist else "No"
            genFont(ws, 'A', row, f'With M.A.R.V.E.L.')
            genFont(ws, 'C', row, f'{marvel_status}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1

            # Add blue header for EXTRACT AIR READINGS
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR READINGS", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Column Headers for CXW
            extendRow(ws, row)
            ws.merge_cells(f'A{row}:B{row}')
            genFont(ws, 'A', row, 'Grill #')
            makeCenter(ws, 'A', row)
            sectionBorder(ws, row, 1, 3)
            
            # Anemometer Reading (m/s)
            genFont(ws, 'C', row, 'Anemometer Reading (m/s)')
            makeCenter(ws, 'C', row)
            sectionBorder(ws, row, 3, 4)
            
            # Free Area (m²)
            genFont(ws, 'D', row, 'Free Area (m²)')
            makeCenter(ws, 'D', row)
            sectionBorder(ws, row, 4, 5)
            
            # Flowrate (m³/h)
            genFont(ws, 'E', row, 'Flowrate (m³/h)')
            makeCenter(ws, 'E', row)
            sectionBorder(ws, row, 5, 6)
            
            # Flowrate (m³/s)
            genFont(ws, 'F', row, 'Flowrate (m³/s)')
            makeCenter(ws, 'F', row)
            sectionBorder(ws, row, 6, 7)
            
            # Only show M.A.R.V.E.L. columns if enabled
            if 'M.A.R.V.E.L. System' in v.checklist:
                # Min %
                genFont(ws, 'G', row, 'Min (%)')
                makeCenter(ws, 'G', row)
                sectionBorder(ws, row, 7, 8)
                
                # Idle %
                genFont(ws, 'H', row, 'Idle (%)')
                makeCenter(ws, 'H', row)
                sectionBorder(ws, row, 8, 9)
                
                # Design m³/s
                genFont(ws, 'I', row, 'Design (m³/s)')
                makeCenter(ws, 'I', row)
                sectionBorder(ws, row, 9, 10)
            else:
                # If M.A.R.V.E.L. is not enabled, show dashes
                genFont(ws, 'G', row, ' -')
                makeCenter(ws, 'G', row)
                sectionBorder(ws, row, 7, 8)
                
                genFont(ws, 'H', row, ' -')
                makeCenter(ws, 'H', row)
                sectionBorder(ws, row, 8, 9)
                
                genFont(ws, 'I', row, ' -')
                makeCenter(ws, 'I', row)
                sectionBorder(ws, row, 9, 10)
            
            row+=1
            
            # Now adjust the data rows to match the new column layout
            total_flowrate_m3s = 0
            for section, info in v.sections.items():
                if isinstance(section, str) and section.isdigit():  # Only process numeric section keys
                    # Merge A and B for section number
                    ws.merge_cells(f'A{row}:B{row}')
                    genFont(ws, 'A', row, section)
                    makeCenter(ws, 'A', row)
                    sectionBorder(ws, row, 1, 3)
                    
                    # Anemometer Reading (m/s)
                    genFont(ws, 'C', row, f' {info["anemometer_reading"]}')
                    makeCenter(ws, 'C', row)
                    sectionBorder(ws, row, 3, 4)
                    
                    # Free Area (m²)
                    genFont(ws, 'D', row, f' {info["free_area"]}')
                    makeCenter(ws, 'D', row)
                    sectionBorder(ws, row, 4, 5)
                    
                    # Flowrate (m³/h)
                    genFont(ws, 'E', row, f' {info["flowrate_m3h"]}')
                    makeCenter(ws, 'E', row)
                    sectionBorder(ws, row, 5, 6)
                    
                    # Flowrate (m³/s)
                    genFont(ws, 'F', row, f' {info["flowrate_m3s"]}')
                    makeCenter(ws, 'F', row)
                    sectionBorder(ws, row, 6, 7)
                    
                    # Add M.A.R.V.E.L. data if enabled
                    if 'M.A.R.V.E.L. System' in v.checklist:
                        # Min %
                        min_percentage = info.get('min_pct', 0.0)
                        genFont(ws, 'G', row, f' {min_percentage}%')
                        makeCenter(ws, 'G', row)
                        sectionBorder(ws, row, 7, 8)
                        
                        # Idle %
                        idle_percentage = info.get('idle_pct', 0.0)
                        genFont(ws, 'H', row, f' {idle_percentage}%')
                        makeCenter(ws, 'H', row)
                        sectionBorder(ws, row, 8, 9)
                        
                        # Design m³/s
                        design_flow = info.get('design_flow', 0.0)
                        genFont(ws, 'I', row, f' {format_flowrate(design_flow, 3)}')
                        makeCenter(ws, 'I', row)
                        sectionBorder(ws, row, 9, 10)
                    else:
                        # If M.A.R.V.E.L. is not enabled, show dashes
                        genFont(ws, 'G', row, ' -')
                        makeCenter(ws, 'G', row)
                        sectionBorder(ws, row, 7, 8)
                        
                        genFont(ws, 'H', row, ' -')
                        makeCenter(ws, 'H', row)
                        sectionBorder(ws, row, 8, 9)
                        
                        genFont(ws, 'I', row, ' -')
                        makeCenter(ws, 'I', row)
                        sectionBorder(ws, row, 9, 10)
                    
                    total_flowrate_m3s += info["flowrate_m3s"]
                    row += 1
            
            # Add blue header before total flowrate 
            colorFill2(ws, row)  # Using colorFill2 which is designed for A:C merging
            row+=1
            
            # Format total with exact 3 decimal places
            formatted_total = format_flowrate(total_flowrate_m3s, 3)
            
            # Total row at the bottom
            ws.merge_cells(f'A{row}:C{row}')  # Merge A to C
            genFont(ws, 'A', row, f'Total Flowrate: {formatted_total} m³/s')
            makeCenter(ws, 'A', row)
            # Only add borders for the merged cell A:C
            sectionBorder(ws, row, 1, 4)
            
            row+=2

        elif v.model in ['KSR-S', 'KSR-F', 'KSR-M']:
            print('lol')
          
        elif v.model in ['CMW-FMOD', 'CMW-IMOD'] and 'slot_length' in v.sections:
            st.write(v.model)  
            # Add EXTRACT AIR DATA header
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR DATA", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Basic hood information
            genFont(ws, 'A', row,f'Drawing Number')
            genFont(ws, 'C', row,f'{v.drawingNum}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Canopy Location')
            genFont(ws, 'C', row, f'{v.location}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Canopy Model')
            genFont(ws, 'C', row, f'{v.model}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Design Flowrate (m³/s)')
            genFont(ws, 'C', row, f'{v.total_design_flow_ms}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Quantity of Sections')
            genFont(ws, 'C', row, f'{v.quantityOfSections}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Length of Slot (mm)')
            genFont(ws, 'C', row, f'{v.sections["slot_length"]}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Width of Slot (mm)')
            genFont(ws, 'C', row, f'{v.sections["slot_width"]}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Calculation')
            genFont(ws, 'C', row, f'Qv = k x √PTAB')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # Add With M.A.R.V.E.L. row for CMW hoods
            marvel_status = "Yes" if "M.A.R.V.E.L. System" in v.checklist else "No"
            genFont(ws, 'A', row, f'With M.A.R.V.E.L. ')
            genFont(ws, 'C', row, f'{marvel_status}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # EXTRACT AIR READINGS header
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR READINGS", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Column headers for reading table
            extendRow(ws, row)
            # Merge A and B for Section #
            ws.merge_cells(f'A{row}:B{row}')
            genFont(ws, 'A', row, 'Section #')
            makeCenter(ws, 'A', row)
            sectionBorder(ws, row, 1, 3)
            
            # TAB Reading - now in column C
            genFont(ws, 'C', row, 'T.A.B Point Reading (Pa)')
            makeCenter(ws, 'C', row)
            sectionBorder(ws, row, 3, 4)
            
            # K-Factor - now in column D
            genFont(ws, 'D', row, 'K-Factor (m³/h)')
            makeCenter(ws, 'D', row)
            sectionBorder(ws, row, 4, 5)
            
            # Flowrate m³/h - now in column E
            genFont(ws, 'E', row, 'Flowrate (m³/h)')
            makeCenter(ws, 'E', row)
            sectionBorder(ws, row, 5, 6)
            
            # Flowrate m³/s - now in column F
            genFont(ws, 'F', row, 'Flowrate (m³/s)')
            makeCenter(ws, 'F', row)
            sectionBorder(ws, row, 6, 7)
            
            # Min % - now in column G
            genFont(ws, 'G', row, 'Min (%)')
            makeCenter(ws, 'G', row)
            sectionBorder(ws, row, 7, 8)
            
            # Idle % - now in column H
            genFont(ws, 'H', row, 'Idle (%)')
            makeCenter(ws, 'H', row)
            sectionBorder(ws, row, 8, 9)
            
            # Design m³/s - now in column I
            genFont(ws, 'I', row, 'Design (m³/s)')
            makeCenter(ws, 'I', row)
            sectionBorder(ws, row, 9, 10)
            
            row+=1
            
            # Now adjust the data rows to match the new column layout
            total_flowrate_m3s = 0
            for section, info in v.sections.items():
                if isinstance(section, str) and section.isdigit():  # Only process numeric section keys
                    # Merge A and B for section number
                    ws.merge_cells(f'A{row}:B{row}')
                    genFont(ws, 'A', row, section)
                    makeCenter(ws, 'A', row)
                    sectionBorder(ws, row, 1, 3)
                    
                    # TAB Reading (Pa)
                    genFont(ws, 'C', row, f' {info["tab_reading"]}')
                    makeCenter(ws, 'C', row)
                    sectionBorder(ws, row, 3, 4)
                    
                    # K-Factor
                    genFont(ws, 'D', row, f' {info["k_factor"]}')
                    makeCenter(ws, 'D', row)
                    sectionBorder(ws, row, 4, 5)
                    
                    # Flowrate (m³/h)
                    genFont(ws, 'E', row, f' {info["flowrate_m3h"]}')
                    makeCenter(ws, 'E', row)
                    sectionBorder(ws, row, 5, 6)
                    
                    # Flowrate (m³/s)
                    genFont(ws, 'F', row, f' {info["flowrate_m3s"]}')
                    makeCenter(ws, 'F', row)
                    sectionBorder(ws, row, 6, 7)
                    
                    total_flowrate_m3s += info["flowrate_m3s"]
                    
                    # Add M.A.R.V.E.L. data if present
                    if 'min_pct' in info and 'M.A.R.V.E.L. System' in v.checklist:
                        genFont(ws, 'G', row, f' {info["min_pct"]}%')
                        makeCenter(ws, 'G', row)
                        sectionBorder(ws, row, 7, 8)
                        
                        genFont(ws, 'H', row, f' {info["idle_pct"]}%')
                        makeCenter(ws, 'H', row)
                        sectionBorder(ws, row, 8, 9)
                        
                        genFont(ws, 'I', row, f' {info["design_flow"]}')
                        makeCenter(ws, 'I', row)
                        sectionBorder(ws, row, 9, 10)
                    else:
                        # If M.A.R.V.E.L. is not enabled, show dashes
                        genFont(ws, 'G', row, ' -')
                        makeCenter(ws, 'G', row)
                        sectionBorder(ws, row, 7, 8)
                        
                        genFont(ws, 'H', row, ' -')
                        makeCenter(ws, 'H', row)
                        sectionBorder(ws, row, 8, 9)
                        
                        genFont(ws, 'I', row, ' -')
                        makeCenter(ws, 'I', row)
                        sectionBorder(ws, row, 9, 10)
                    
                    row += 1
            
            # Format total with exact 3 decimal places
            formatted_total = format_flowrate(total_flowrate_m3s, 3)
            
            # Add blue box before total flowrate row
            colorFill2(ws, row)
            row += 1
            
            # Total row at the bottom
            ws.merge_cells(f'A{row}:C{row}')  # Merge A to C
            genFont(ws, 'A', row, f'Total Flowrate: {formatted_total} m³/s')
            makeCenter(ws, 'A', row)
            # Only add borders for the merged cell A:C
            sectionBorder(ws, row, 1, 4)
            
            row+=2

        elif v.model in ['UVF', 'UVI']:
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR DATA", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1
            
            # Basic hood information in CXW style format
            genFont(ws, 'A', row, f'Drawing Number')
            genFont(ws, 'C', row, f'{v.drawingNum}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Location')
            genFont(ws, 'C', row, f'{v.location}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Model')
            genFont(ws, 'C', row, f'{v.model}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Design Flowrate (m³/s)')
            genFont(ws, 'C', row, f'{v.total_design_flow_ms}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            genFont(ws, 'A', row, f'Quantity of Sections')
            genFont(ws, 'C', row, f'{v.quantityOfSections}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # Add With M.A.R.V.E.L. row
            marvel_status = "Yes" if "M.A.R.V.E.L. System" in v.checklist else "No"
            genFont(ws, 'A', row, f'With M.A.R.V.E.L.')
            genFont(ws, 'C', row, f'{marvel_status}')
            ws.merge_cells(f'A{row}:B{row}')
            ws.merge_cells(f'C{row}:I{row}')
            sectionBorder(ws, row, 1, 3)
            sectionBorder(ws, row, 3, 10)
            row+=1
            
            # Add section readings header
            colorFill(ws, row)
            ws.merge_cells(f'A{row}:I{row}')
            genFont(ws, 'A', row, "EXTRACT AIR READINGS", "FFFFFF")
            makeCenter(ws, 'A', row)
            row+=1

        row+=3
        #Results
       # Results Summary - Extract Air
         # Results Summary for Extract Air
    extract_summary = []
    supply_summary = []
    total_extract_design = 0
    total_extract_actual = 0
    total_supply_design = 0
    total_supply_actual = 0

    # Iterate over each canopy to process both Extract and Supply Air
    for canopy_num, hood in hoods.items():
        drawing_number = f"{hood.drawingNum}"  # Format the drawing number

        # Initialize per-drawing totals
        extract_design_total = 0
        extract_actual_total = 0
        supply_design_total = 0
        supply_actual_total = 0

        # Process Extract Air sections
        # st.write(hood)
        if hood.model in ['KVF', 'KVI', 'KCH-F', 'KCH-I', 'KSR-S', 'KSR-F', 'KSR-M', 'UVF', 'UVI', 'USR-S', 'USR-F', 'USR-M', 'UWF'] or (hood.model == 'CMW-FMOD' and 'slot_length' not in hood.sections):
            for section_data in hood.sections.values():
                if isinstance(section_data, dict) and "k_factor" in section_data and "tab_reading" in section_data:
                    extract_design_total = hood.total_design_flow_ms  # Use the total extract design flow
                    # Calculate actual flow - store the raw value to avoid precision loss
                    flow_value = (section_data["k_factor"] * math.sqrt(section_data["tab_reading"])) / 3600
                    extract_actual_total += flow_value  # Store raw value, not formatted

            # Calculate percentage using total values
            if extract_design_total > 0:
                extract_percentage = format_flowrate((extract_actual_total / extract_design_total) * 100, 1)
        
        # Process CXW hoods differently since they use a different calculation method
        
        elif hood.model == 'CXW':
            extract_design_total = hood.total_design_flow_ms
            extract_actual_total = hood.total_flowrate_m3s  # This should already be the raw value
            
            # Calculate percentage
            if extract_design_total > 0:
                extract_percentage = format_flowrate((extract_actual_total / extract_design_total) * 100, 1)
                
        # Process CMW hoods with slot dimensions
        elif hood.model in ['CMW-FMOD', 'CMW-IMOD'] and 'slot_length' in hood.sections:
            extract_design_total = hood.total_design_flow_ms
            extract_actual_total = 0
            
            # Sum up the flowrates from each section
            for section_key, section_data in hood.sections.items():
                # Only process numeric section keys (skip 'slot_length', 'slot_width')
                if isinstance(section_key, str) and section_key.isdigit() and isinstance(section_data, dict) and "flowrate_m3s" in section_data:
                    extract_actual_total += section_data["flowrate_m3s"]
            
            # Calculate percentage
            if extract_design_total > 0:
                extract_percentage = format_flowrate((extract_actual_total / extract_design_total) * 100, 1)

        # Process Supply Air sections if applicable
        if hood.model in ['KVF', 'KCH-F', 'UVF', 'USR-F', 'KSR-F', 'KWF', 'UWF'] or (hood.model == 'CMW-FMOD' and 'slot_length' not in hood.sections):
            for section_data in hood.sections.values():
                if isinstance(section_data, dict) and "supplyKFactor" in section_data and "supplyTab" in section_data:
                    supply_design_total = hood.total_supply_design_flow_ms  # Use the total design flow for whole canopy
                    # Calculate actual flow - store the raw value to avoid precision loss
                    flow_value = (section_data["supplyKFactor"] * math.sqrt(section_data["supplyTab"])) / 3600
                    supply_actual_total += flow_value  # Store raw value, not formatted

            # Calculate percentage using total values
            if supply_design_total > 0:
                supply_percentage = format_flowrate((supply_actual_total / supply_design_total) * 100, 1)

        # Calculate percentage and store results for each drawing - store both raw and formatted values
        if extract_design_total > 0:
            # Format for display, but keep raw value for later calculations
            formatted_extract_actual = format_flowrate(extract_actual_total, 2)  # Use 2 decimal places
            extract_summary.append({
                "drawing_number": hood.drawingNum,
                "design_flow_rate_total": extract_design_total,
                "actual_flow_rate_total": formatted_extract_actual,  # Formatted for display
                "raw_actual_flow_rate": extract_actual_total,  # Raw for calculations
                "percentage": extract_percentage
            })
            total_extract_design += extract_design_total
            total_extract_actual += extract_actual_total  # Use raw value

        if supply_design_total > 0:
            # Format for display, but keep raw value for later calculations
            formatted_supply_actual = format_flowrate(supply_actual_total, 2)  # Use 2 decimal places
            supply_summary.append({
                "drawing_number": drawing_number,
                "design_flow_rate_total": supply_design_total,
                "actual_flow_rate_total": formatted_supply_actual,  # Formatted for display
                "raw_actual_flow_rate": supply_actual_total,  # Raw for calculations
                "percentage": supply_percentage
            })
            total_supply_design += supply_design_total
            total_supply_actual += supply_actual_total  # Use raw value

    # Display Results Summary for Extract Air
    row += 2
    genFont(ws, 'A', row, "RESULTS SUMMARY - EXTRACT AIR")
    row += 1
    colorFill(ws, row)

    # Header for Extract Air Summary Table
    ws.merge_cells(f'A{row}:B{row}')
    genFont(ws, 'A', row, "Drawing Number", "FFFFFF")
    makeCenter(ws, 'A', row)
    
    ws.merge_cells(f'C{row}:D{row}')
    genFont(ws, 'C', row, "Design Flow Rate (m³/s)", "FFFFFF")
    makeCenter(ws, 'C', row)
    
    ws.merge_cells(f'E{row}:F{row}')
    genFont(ws, 'E', row, "Actual Flowrate (m³/s)", "FFFFFF")
    makeCenter(ws, 'E', row)
    
    ws.merge_cells(f'G{row}:I{row}')
    genFont(ws, 'G', row, "Percentage of Design", "FFFFFF")
    makeCenter(ws, 'G', row)
    
    sectionBorder(ws, row, 1, 10)
    row += 1

    # Initialize totals
    total_design_flow = 0
    total_actual_flow = 0

    # Process each hood
    for hood_key, hood in hoods.items():
        if hood.model in ['KVF', 'KVI', 'KCH-F', 'KCH-I', 'KSR-S', 'KSR-F', 'KSR-M', 'UVF', 'UVI', 'USR-S', 'USR-F', 'USR-M', 'CXW'] or (hood.model in ['CMW-FMOD', 'CMW-IMOD']):
            # Calculate actual flow rate for this hood
            actual_flow = 0
            for section, info in hood.sections.items():
                if isinstance(info, dict):
                    # This works
                    if hood.model == 'CXW':
                        if 'flowrate_m3s' in info:
                            actual_flow += info['flowrate_m3s']
                    elif hood.model in ['CMW-FMOD', 'CMW-IMOD']:
                        # Use k-factor method for CMW hoods
                        if 'k_factor' in info and 'tab_reading' in info:
                            section_flow = (info['k_factor'] * math.sqrt(info['tab_reading'])) / 3600
                            actual_flow += section_flow
                    else:
                        if 'k_factor' in info and 'tab_reading' in info:
                            section_flow = (info['k_factor'] * math.sqrt(info['tab_reading'])) / 3600
                            actual_flow += section_flow

            # Format values
            design_flow = hood.total_design_flow_ms
            actual_flow = round(actual_flow, 3)
            percentage = round((actual_flow / design_flow * 100), 1) if design_flow > 0 else 0

            # Add to totals
            total_design_flow += design_flow
            total_actual_flow += actual_flow

            # Display row
            ws.merge_cells(f'A{row}:B{row}')
            genFont(ws, 'A', row, str(hood.drawingNum))
            makeCenter(ws, 'A', row)
            
            ws.merge_cells(f'C{row}:D{row}')
            genFont(ws, 'C', row, f"{format_flowrate(design_flow, 3)}")
            makeCenter(ws, 'C', row)
            
            ws.merge_cells(f'E{row}:F{row}')
            genFont(ws, 'E', row, f"{format_flowrate(actual_flow, 3)}")
            makeCenter(ws, 'E', row)
            
            ws.merge_cells(f'G{row}:I{row}')
            genFont(ws, 'G', row, f"{percentage}%")
            makeCenter(ws, 'G', row)
            
            sectionBorder(ws, row, 1, 10)
            row += 1

    # Display totals row
    ws.merge_cells(f'A{row}:B{row}')
    genFont(ws, 'A', row, "TOTAL")
    makeCenter(ws, 'A', row)
    
    ws.merge_cells(f'C{row}:D{row}')
    genFont(ws, 'C', row, f"{format_flowrate(total_design_flow, 3)}")
    makeCenter(ws, 'C', row)
    
    ws.merge_cells(f'E{row}:F{row}')
    genFont(ws, 'E', row, f"{format_flowrate(total_actual_flow, 3)}")
    makeCenter(ws, 'E', row)
    
    ws.merge_cells(f'G{row}:I{row}')
    total_percentage = round((total_actual_flow / total_design_flow * 100), 1) if total_design_flow > 0 else 0
    genFont(ws, 'G', row, f"{total_percentage}%")
    makeCenter(ws, 'G', row)
    
    sectionBorder(ws, row, 1, 10)
    row += 3

    # Display Results Summary for Supply Air
    genFont(ws, 'A', row, "RESULTS SUMMARY - SUPPLY AIR")
    row += 1
    colorFill(ws, row)

    # Header for Supply Air Summary Table
    ws.merge_cells(f'A{row}:B{row}')
    genFont(ws, 'A', row, "Drawing Number", "FFFFFF")
    makeCenter(ws, 'A', row)
    
    ws.merge_cells(f'C{row}:D{row}')
    genFont(ws, 'C', row, "Design Flow Rate (m³/s)", "FFFFFF")
    makeCenter(ws, 'C', row)
    
    ws.merge_cells(f'E{row}:F{row}')
    genFont(ws, 'E', row, "Actual Flowrate (m³/s)", "FFFFFF")
    makeCenter(ws, 'E', row)
    
    ws.merge_cells(f'G{row}:I{row}')
    genFont(ws, 'G', row, "Percentage of Design", "FFFFFF")
    makeCenter(ws, 'G', row)
    
    sectionBorder(ws, row, 1, 10)
    row += 1

    # Initialize supply totals
    total_supply_design = 0.0
    total_supply_actual = 0.0

    # Process each hood for supply air
    for hood_key, hood in hoods.items():
        if hood.model in ['KVF', 'KCH-F', 'UVF', 'USR-F', 'KSR-F', 'KWF', 'UWF'] or (hood.model == 'CMW-FMOD'):
            # Calculate actual supply flow rate for this hood
            actual_supply = 0.0
            for section, info in hood.sections.items():
                if isinstance(info, dict):
                    if 'supplyKFactor' in info and 'supplyTab' in info:
                        section_flow = (info['supplyKFactor'] * math.sqrt(info['supplyTab'])) / 3600
                        actual_supply += section_flow

            # Format values
            design_supply = hood.total_supply_design_flow_ms
            actual_supply = round(actual_supply, 3)
            supply_percentage = round((actual_supply / design_supply * 100), 1) if design_supply > 0 else 0

            # Add to totals
            total_supply_design += design_supply
            total_supply_actual += actual_supply

            # Display row
            ws.merge_cells(f'A{row}:B{row}')
            genFont(ws, 'A', row, str(hood.drawingNum))
            makeCenter(ws, 'A', row)
            
            ws.merge_cells(f'C{row}:D{row}')
            genFont(ws, 'C', row, f"{format_flowrate(design_supply, 3)}")
            makeCenter(ws, 'C', row)
            
            ws.merge_cells(f'E{row}:F{row}')
            genFont(ws, 'E', row, f"{format_flowrate(actual_supply, 3)}")
            makeCenter(ws, 'E', row)
            
            ws.merge_cells(f'G{row}:I{row}')
            genFont(ws, 'G', row, f"{supply_percentage}%")
            makeCenter(ws, 'G', row)
            
            sectionBorder(ws, row, 1, 10)
            row += 1

    # Display supply totals row
    ws.merge_cells(f'A{row}:B{row}')
    genFont(ws, 'A', row, "TOTAL")
    makeCenter(ws, 'A', row)
    
    ws.merge_cells(f'C{row}:D{row}')
    genFont(ws, 'C', row, f"{format_flowrate(total_supply_design, 3)}")
    makeCenter(ws, 'C', row)
    
    ws.merge_cells(f'E{row}:F{row}')
    genFont(ws, 'E', row, f"{format_flowrate(total_supply_actual, 3)}")
    makeCenter(ws, 'E', row)
    
    ws.merge_cells(f'G{row}:I{row}')
    total_supply_percentage = round((total_supply_actual / total_supply_design * 100), 1) if total_supply_design > 0 else 0
    genFont(ws, 'G', row, f"{total_supply_percentage}%")
    makeCenter(ws, 'G', row)
    
    sectionBorder(ws, row, 1, 10)
    row += 3

    if edge_box_details:
    # Add header row with blue fill
        row += 5
        genFont(ws, 'A', row, "FINAL SYSTEM CHECKS")
        row += 1
        colorFill(ws, row)  # Blue background
        ws.merge_cells(f'A{row}:G{row}')
        ws.merge_cells(f'H{row}:I{row}')
        genFont(ws, 'A', row, "EDGE BOX", "FFFFFF")
        makeCenter(ws, 'A', row)

        # Add Edge Box parameters
        row += 1
        for key, value in edge_box_details.items():
            # Merge columns A:G for the parameter name
            ws.merge_cells(f'A{row}:G{row}')
            genFont(ws, 'A', row, key)

            # Merge columns H:I for the parameter value
            ws.merge_cells(f'H{row}:I{row}')
            if value is True:
                ws[f'H{row}'] = "✓"
            elif value is False:
                ws[f'H{row}'] = ""
            else:
                ws[f'H{row}'] = value  # For non-boolean values

            # Center-align the value
            makeCenter(ws, 'H', row)

            # Add borders for the row
            sectionBorder(ws, row, 1, 10)
            row += 1
                
                
    #Additional Notes
    row+=5
    colorFill(ws, row)
    ws.merge_cells('A{0}:I{0}'.format(row))
    genFont(ws, 'A', row, f'ADDITIONAL NOTES', "FFFFFF")
    row+=1
    for comment in comments:
        ws[f'A{row}'] = comment
        extendRow(ws, row)
        ws.merge_cells(f'A{row}:I{row}')
        makeCenter(ws, 'A', row)
        ws[f'A{row}'].border = Border(left=Side(style='thin'))
        ws[f'I{row}'].border = Border(right=Side(style='thin'))
        row+=1
    ws[f'A{row}'].border = Border(top=Side(style='thin'))
    ws[f'B{row}'].border = Border(top=Side(style='thin'))
    ws[f'C{row}'].border = Border(top=Side(style='thin'))
    ws[f'D{row}'].border = Border(top=Side(style='thin'))
    ws[f'E{row}'].border = Border(top=Side(style='thin'))
    ws[f'F{row}'].border = Border(top=Side(style='thin'))
    ws[f'G{row}'].border = Border(top=Side(style='thin'))
    ws[f'H{row}'].border = Border(top=Side(style='thin'))
    ws[f'I{row}'].border = Border(top=Side(style='thin'))
    genFont(ws, 'A', row, "Date")
    genFont(ws, 'C', row, f"{genInfo['date_of_visit']}")
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:I{row}')
    sectionBorder(ws, row, 1, 3)
    sectionBorder(ws, row, 3, 10)
    row+=1
    genFont(ws, 'A', row, "Print")
    genFont(ws, 'C', row, f"{genInfo['engineers']}")
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:I{row}')
    sectionBorder(ws, row, 1, 3)
    sectionBorder(ws, row, 3, 10)
    row += 1

    # Only add signature if it exists
    if sign.image_data is not None:
        # Add "Sign" row
        genFont(ws, 'A', row, "Sign")
        ws.merge_cells(f'A{row}:I{row}')
        # Apply only left and right borders to the merged range
        ws[f'A{row}'].border = Border(left=Side(style='thin'))
        ws[f'I{row}'].border = Border(right=Side(style='thin'))
        row += 1
        
        start_row = row
        # Create 6 vertical rows for signature
        for i in range(6):
            ws.merge_cells(f'A{start_row + i}:I{start_row + i}')
            # Add left and right borders
            ws[f'A{start_row + i}'].border = Border(left=Side(style='thin'))
            ws[f'I{start_row + i}'].border = Border(right=Side(style='thin'))
        
        # Add bottom, left, and right borders to the last row
        ws[f'A{start_row + 5}'].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'I{start_row + 5}'].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{start_row + 5}'].border = Border(bottom=Side(style='thin'))
        
        # Add the signature image with adjusted size and position
        img = PILImage.fromarray(sign.image_data.astype('uint8'), 'RGBA')
        # Step 1: Resize the image (change dimensions as needed, e.g., 150x75 pixels)
        resized_img = img.resize((200, 75))  # Resize to desired dimensions
        
        # Step 2: Save the resized image in-memory and to a file
        img_io = io.BytesIO()
        resized_img.save(img_io, format='PNG')
        img_io.seek(0)
        
        with open("resized_signature_image.png", "wb") as f:
            f.write(img_io.read())
        
        # Step 3: Load the resized image into openpyxl
        signature_img = Image("resized_signature_image.png")
        
        # Step 4: Insert the signature into the Excel file
        ws.add_image(signature_img, f"D{start_row + 1}")  # Moved to column D and down one row
        
        row += 6

            
        ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=row))  # Break after Row 20 (before row 21)
            #End of Capture Jet Hoods

        row+=2
        
    

    # You can also add more rows if necessary, e.g., canopy details, etc.
    #applyMulishFont(ws)
    # Define the path for the new Excel file
    updated_workbook_path = f'T&C.xlsx'
    
    # Save the workbook to the specified path
    wb.save(updated_workbook_path)
    
    # Provide the download link
    with open(updated_workbook_path, "rb") as file:
        st.download_button(
            label="Download Report", 
            data=file, 
            file_name=f"{genInfo['project_number']} {genInfo['project_name']} - Canopy Commissioning Report.xlsx"
        )
      
def iBorder(ws, letter, row):
    ws[f'i{row}'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws[f'A{row}'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
def genFont(ws, col, row, text, color="000000"):  # Added color parameter with default black
    ws[f'{col}{row}'] = text
    ws[f'{col}{row}'].font = Font(name='Calibri', size=11, color=color)
def extendRow(ws, row):
    ws.row_dimensions[row].height = 30
def makeCenter(ws, letter, row):
    ws[f'{letter}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
def sectionBorder(ws,row, first, second):
    thin_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
    )
    for col in range(first, second):  # Columns A to D
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        
def titleFont(ws, letter, row, word):
    ws[f'{letter}{row}'] = word
    ws[f'{letter}{row}'].font = Font(size=14, bold=True)
    
def fillBorder(ws, letter, row):
    thin_border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
    for col in range(1, 10):  # Column A to B (1 to 2)
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Apply border
            
def colorFill(ws, row):
    fill_style = PatternFill(start_color="2499D5", end_color="2499D5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define the start and end columns for the merged range
    start_col = 1  # Column A
    end_col = 9    # Column I

    # Loop through each column in the merged range
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_style
        cell.border = thin_border

    # Ensure the left and right borders for the first and last columns
    ws.cell(row=row, column=start_col).border = Border(
        left=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.cell(row=row, column=end_col).border = Border(
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def colorFill2(ws, row):
    # Change the merging from A:B to A:C
    ws.merge_cells(f'A{row}:C{row}')
    
    # Apply fill to columns A, B, and C
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].fill = PatternFill(start_color="2499D5", end_color="2499D5", fill_type="solid")
    
    # Add borders only on the left of A and right of C
    ws[f'A{row}'].border = Border(left=Side(style='thin'))
    ws[f'C{row}'].border = Border(right=Side(style='thin'))

def applyMulishFont(ws):
    """
    Apply Mulish font to all cells in the worksheet.
    """
    mulish_font = Font(name='Arial', size=11)  # Define the Mulish font

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:  # Apply font only if the cell contains content
                cell.font = mulish_font

            print("Current row:", row)
            print("Merged cells before clearing:", [str(r) for r in ws.merged_cells.ranges])
            
            # Clear any existing merges in this row
            for merge_range in ws.merged_cells.ranges.copy():
                if row in range(merge_range.min_row, merge_range.max_row + 1):
                    print(f"Unmerging: {str(merge_range)}")
                    ws.unmerge_cells(str(merge_range))
            
            print("Attempting to merge G to I:", f"G{row}:I{row}")
            ws.merge_cells(f'G{row}:I{row}')
            print("Merged cells after operation:", [str(r) for r in ws.merged_cells.ranges])
            
            genFont(ws, 'G', row, 'Percentage of Design', "FFFFFF")
            makeCenter(ws, 'G', row)